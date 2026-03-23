from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Aiden AB Test v2"

header_fill = PatternFill('solid', fgColor='2C3E50')
bean_fill = PatternFill('solid', fgColor='1ABC9C')
warn_fill = PatternFill('solid', fgColor='FADBD8')
pass_fill = PatternFill('solid', fgColor='D5F5E3')
white = PatternFill('solid', fgColor='FFFFFF')
alt_row = PatternFill('solid', fgColor='F8F9FA')
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
bean_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
label_font = Font(name='Arial', bold=True, size=10)
data_font = Font(name='Arial', size=10)
match_font = Font(name='Arial', size=10, color='27AE60', bold=True)
thin_border = Border(
    left=Side(style='thin', color='D5D8DC'), right=Side(style='thin', color='D5D8DC'),
    top=Side(style='thin', color='D5D8DC'), bottom=Side(style='thin', color='D5D8DC'),
)

beans = [
    {
        'name': 'El Placer',
        'desc': 'Dayglow | Colombia | Geisha | Anaerobic White Honey | In Peak 67%',
        'target_family': 'PROCESSED CLARITY',
        'target_grind': '~4.4',
        'gpt_orig': {'title':'Promethium Tea Time at 7','ratio':17,'bloomRatio':2.5,'bloomDuration':'45s','bloomTemp':'92°C','ssPulses':3,'ssInterval':'25s','ssTemps':'92.5, 92, 91.5','batchPulses':4,'batchInterval':'30s','batchTemps':'92.5, 92, 91.5, 91.5','ssGrind':4.4,'batchGrind':6.4,'validation':'PASS'},
        'sonnet_old': {'title':'El Placer: Gardenia & Earl Grey','ratio':16.5,'bloomRatio':3,'bloomDuration':'45s','bloomTemp':'90°C','ssPulses':3,'ssInterval':'28s','ssTemps':'93, 92.5, 91.5','batchPulses':3,'batchInterval':'33s','batchTemps':'93, 92.5, 91.5','ssGrind':5.1,'batchGrind':6.2,'validation':'PASS'},
        'sonnet_new': {'title':'El Placer - Gardenia & Earl Grey','ratio':17.5,'bloomRatio':2.5,'bloomDuration':'50s','bloomTemp':'93°C','ssPulses':3,'ssInterval':'27s','ssTemps':'93, 92, 91','batchPulses':3,'batchInterval':'30s','batchTemps':'93, 92, 91','ssGrind':5.2,'batchGrind':7,'validation':'PASS'},
        'gpt54_new': {'title':'El Placer Bergamot Glow','ratio':17.5,'bloomRatio':2.5,'bloomDuration':'50s','bloomTemp':'93°C','ssPulses':3,'ssInterval':'28s','ssTemps':'93, 92.5, 92','batchPulses':3,'batchInterval':'30s','batchTemps':'93, 92.5, 92','ssGrind':5.2,'batchGrind':6.2,'validation':'PASS'},
    },
    {
        'name': 'Finca La Fuente',
        'desc': 'Koppi | Colombia | Pink Bourbon | Washed | In Peak 82%',
        'target_family': 'WASHED FLORAL CLARITY',
        'target_grind': '~4.2',
        'gpt_orig': {'title':'Koppi La Fuente Pink Bloom','ratio':17,'bloomRatio':3,'bloomDuration':'45s','bloomTemp':'95.5°C','ssPulses':3,'ssInterval':'22s','ssTemps':'96, 95, 94','batchPulses':4,'batchInterval':'28s','batchTemps':'96, 95, 94, 93','ssGrind':4.2,'batchGrind':6.2,'validation':'PASS'},
        'sonnet_old': {'title':'Koppi La Fuente Pink Bloom','ratio':17,'bloomRatio':3,'bloomDuration':'45s','bloomTemp':'96°C','ssPulses':3,'ssInterval':'20s','ssTemps':'96, 95, 94','batchPulses':3,'batchInterval':'25s','batchTemps':'96, 95, 94','ssGrind':5.2,'batchGrind':7,'validation':'PASS'},
        'sonnet_new': {'title':'La Fuente Pink Bloom','ratio':17.5,'bloomRatio':3,'bloomDuration':'50s','bloomTemp':'95°C','ssPulses':3,'ssInterval':'23s','ssTemps':'95, 94, 93','batchPulses':3,'batchInterval':'25s','batchTemps':'95, 94, 93','ssGrind':6,'batchGrind':7,'validation':'PASS'},
        'gpt54_new': {'title':'La Fuente Floral Lift','ratio':17.5,'bloomRatio':3,'bloomDuration':'50s','bloomTemp':'95°C','ssPulses':3,'ssInterval':'23s','ssTemps':'95, 94, 93','batchPulses':3,'batchInterval':'25s','batchTemps':'95, 94, 93','ssGrind':5.2,'batchGrind':6.2,'validation':'PASS'},
    },
    {
        'name': 'Mulish',
        'desc': "Apollon's Gold | Ethiopia | Heirloom | Washed | Past Peak +15d",
        'target_family': 'WASHED ETHIOPIA CLARITY',
        'target_grind': '~4.2',
        'gpt_orig': {'title':"Apollon's Gold Mulish",'ratio':17,'bloomRatio':3,'bloomDuration':'50s','bloomTemp':'94.5°C','ssPulses':3,'ssInterval':'22s','ssTemps':'95.5, 94.5, 93.5','batchPulses':4,'batchInterval':'28s','batchTemps':'95.5, 94.5, 93.5, 93','ssGrind':4.2,'batchGrind':6.2,'validation':'PASS'},
        'sonnet_old': {'title':'Mulish - Faded Nectarine','ratio':17.5,'bloomRatio':3.5,'bloomDuration':'45s','bloomTemp':'99°C','ssPulses':3,'ssInterval':'20s','ssTemps':'99, 99, 98','batchPulses':4,'batchInterval':'25s','batchTemps':'99, 99, 98, 97','ssGrind':5.1,'batchGrind':7,'validation':'bloomRatio > 3 ⚠'},
        'sonnet_new': {'title':'Mulish - Fading Bloom','ratio':17.5,'bloomRatio':3.5,'bloomDuration':'50s','bloomTemp':'96.5°C','ssPulses':3,'ssInterval':'22s','ssTemps':'96.5, 95.5, 94.5','batchPulses':3,'batchInterval':'23s','batchTemps':'96.5, 95.5, 94.5','ssGrind':5.2,'batchGrind':7,'validation':'bloomRatio > 3 ⚠'},
        'gpt54_new': {'title':'Mulish Lavender Lift','ratio':17.5,'bloomRatio':3.5,'bloomDuration':'50s','bloomTemp':'96°C','ssPulses':3,'ssInterval':'22s','ssTemps':'96, 95, 94','batchPulses':4,'batchInterval':'30s','batchTemps':'96, 95, 94, 93','ssGrind':5.2,'batchGrind':6.2,'validation':'bloomRatio > 3 ⚠'},
    },
]

fields = [
    ('Title', 'title'), ('Ratio', 'ratio'), ('Bloom Ratio', 'bloomRatio'),
    ('Bloom Duration', 'bloomDuration'), ('Bloom Temp', 'bloomTemp'),
    ('SS Pulses', 'ssPulses'), ('SS Interval', 'ssInterval'), ('SS Temps', 'ssTemps'),
    ('Batch Pulses', 'batchPulses'), ('Batch Interval', 'batchInterval'), ('Batch Temps', 'batchTemps'),
    ('SS Grind (Ode)', 'ssGrind'), ('Batch Grind (Ode)', 'batchGrind'),
    ('Validation', 'validation'),
]

ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 28
ws.column_dimensions['D'].width = 28
ws.column_dimensions['E'].width = 28

row = 1
ws.merge_cells('A1:E1')
ws['A1'] = 'Aiden Recipe A/B Test v2: Updated Family-First Prompt'
ws['A1'].font = Font(name='Arial', bold=True, size=14)
ws['A1'].alignment = Alignment(horizontal='center')

ws.merge_cells('A2:E2')
ws['A2'] = 'Comparing: ChatGPT Baseline | Old Prompt (Sonnet) | New Prompt (Sonnet 4.6) | New Prompt (GPT-5.4)'
ws['A2'].font = Font(name='Arial', italic=True, size=10, color='666666')
ws['A2'].alignment = Alignment(horizontal='center')
row = 4

for bean in beans:
    # Bean header
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws.cell(row=row, column=1, value=f"{bean['name']} — {bean['desc']}")
    cell.font = bean_font; cell.fill = bean_fill; cell.alignment = Alignment(horizontal='center')
    for c in range(1, 6):
        ws.cell(row=row, column=c).fill = bean_fill
        ws.cell(row=row, column=c).border = thin_border
    row += 1

    # Target info
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1, value=f"Target Family: {bean['target_family']}  |  Target SS Grind: {bean['target_grind']}")
    ws.cell(row=row, column=1).font = Font(name='Arial', bold=True, size=10, color='8E44AD')
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 1

    # Model headers
    headers = ['Parameter', 'ChatGPT Baseline', 'Old Prompt (Sonnet)', 'New Prompt (Sonnet 4.6)', 'New Prompt (GPT-5.4)']
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i+1, value=h)
        c.font = header_font; c.fill = header_fill; c.border = thin_border
        c.alignment = Alignment(horizontal='center')
    row += 1

    models = ['gpt_orig', 'sonnet_old', 'sonnet_new', 'gpt54_new']
    for fi, (label, key) in enumerate(fields):
        bg = alt_row if fi % 2 == 0 else white
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=1).fill = bg; ws.cell(row=row, column=1).border = thin_border

        for ci, mk in enumerate(models):
            val = bean[mk][key]
            cell = ws.cell(row=row, column=ci+2, value=val)
            cell.font = data_font; cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if key == 'validation':
                cell.fill = pass_fill if val == 'PASS' else warn_fill
                cell.font = Font(name='Arial', bold=True, size=10)
            elif key == 'bloomRatio' and isinstance(val, (int, float)) and val > 3:
                cell.fill = warn_fill
                cell.font = Font(name='Arial', bold=True, size=10, color='C0392B')
            else:
                cell.fill = bg
        row += 1
    row += 1

wb.save('/Users/talmeltzer/Documents/VIBE CODING/Coffee App Build /docs/data/aiden-ab-test-results-v2.xlsx')
print('Saved docs/data/aiden-ab-test-results-v2.xlsx')
