from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Aiden AB Test Results"

# Colors
header_fill = PatternFill('solid', fgColor='2C3E50')
bean_fill = PatternFill('solid', fgColor='1ABC9C')
sonnet_fill = PatternFill('solid', fgColor='D6EAF8')
gpt_fill = PatternFill('solid', fgColor='D5F5E3')
warn_fill = PatternFill('solid', fgColor='FADBD8')
pass_fill = PatternFill('solid', fgColor='D5F5E3')
white = PatternFill('solid', fgColor='FFFFFF')
alt_row = PatternFill('solid', fgColor='F8F9FA')
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
bean_font = Font(name='Arial', bold=True, color='FFFFFF', size=12)
label_font = Font(name='Arial', bold=True, size=10)
data_font = Font(name='Arial', size=10)
thin_border = Border(
    left=Side(style='thin', color='D5D8DC'),
    right=Side(style='thin', color='D5D8DC'),
    top=Side(style='thin', color='D5D8DC'),
    bottom=Side(style='thin', color='D5D8DC'),
)

beans = [
    {
        'name': 'El Placer',
        'desc': 'Dayglow | Colombia | Geisha | Anaerobic White Honey | In Peak 67%',
        'sonnet': {
            'title': 'El Placer: Gardenia & Earl Grey',
            'ratio': 16.5, 'bloomRatio': 3, 'bloomDuration': '45s', 'bloomTemp': '90°C',
            'ssPulses': 3, 'ssInterval': '28s', 'ssTemps': '93, 92.5, 91.5',
            'batchPulses': 3, 'batchInterval': '33s', 'batchTemps': '93, 92.5, 91.5',
            'ssGrind': 5.1, 'batchGrind': 6.2, 'validation': 'PASS', 'latency': '3904ms',
        },
        'gpt': {
            'title': 'El Placer Bergamot Bloom',
            'ratio': 16.5, 'bloomRatio': 3, 'bloomDuration': '45s', 'bloomTemp': '92°C',
            'ssPulses': 3, 'ssInterval': '28s', 'ssTemps': '92.5, 92, 91',
            'batchPulses': 3, 'batchInterval': '32s', 'batchTemps': '92.5, 92, 91',
            'ssGrind': 5.2, 'batchGrind': 6.2, 'validation': 'PASS', 'latency': '2915ms',
        },
    },
    {
        'name': 'Finca La Fuente',
        'desc': 'Koppi | Colombia | Pink Bourbon | Washed | In Peak 82%',
        'sonnet': {
            'title': 'Koppi La Fuente Pink Bloom',
            'ratio': 17, 'bloomRatio': 3, 'bloomDuration': '45s', 'bloomTemp': '96°C',
            'ssPulses': 3, 'ssInterval': '20s', 'ssTemps': '96, 95, 94',
            'batchPulses': 3, 'batchInterval': '25s', 'batchTemps': '96, 95, 94',
            'ssGrind': 5.2, 'batchGrind': 7, 'validation': 'PASS', 'latency': '4407ms',
        },
        'gpt': {
            'title': 'La Fuente Pink Bloom',
            'ratio': 17, 'bloomRatio': 2.5, 'bloomDuration': '45s', 'bloomTemp': '97°C',
            'ssPulses': 3, 'ssInterval': '22s', 'ssTemps': '97, 95, 94',
            'batchPulses': 3, 'batchInterval': '25s', 'batchTemps': '97, 95, 94',
            'ssGrind': 5.2, 'batchGrind': 6.2, 'validation': 'PASS', 'latency': '2968ms',
        },
    },
    {
        'name': 'Mulish',
        'desc': "Apollon's Gold | Ethiopia | Heirloom | Washed | Past Peak +15d",
        'sonnet': {
            'title': 'Mulish - Faded Nectarine',
            'ratio': 17.5, 'bloomRatio': 3.5, 'bloomDuration': '45s', 'bloomTemp': '99°C',
            'ssPulses': 3, 'ssInterval': '20s', 'ssTemps': '99, 99, 98',
            'batchPulses': 4, 'batchInterval': '25s', 'batchTemps': '99, 99, 98, 97',
            'ssGrind': 5.1, 'batchGrind': 7, 'validation': 'bloomRatio > 3 ⚠', 'latency': '2929ms',
        },
        'gpt': {
            'title': 'Mulish Floral Lift',
            'ratio': 17, 'bloomRatio': 3.5, 'bloomDuration': '45s', 'bloomTemp': '99°C',
            'ssPulses': 3, 'ssInterval': '20s', 'ssTemps': '99, 98, 97',
            'batchPulses': 4, 'batchInterval': '25s', 'batchTemps': '99, 98, 96, 95',
            'ssGrind': 5.2, 'batchGrind': 7.1, 'validation': 'bloomRatio > 3 ⚠', 'latency': '3027ms',
        },
    },
]

fields = [
    ('Title', 'title'), ('Ratio', 'ratio'), ('Bloom Ratio', 'bloomRatio'),
    ('Bloom Duration', 'bloomDuration'), ('Bloom Temp', 'bloomTemp'),
    ('SS Pulses', 'ssPulses'), ('SS Interval', 'ssInterval'), ('SS Temps', 'ssTemps'),
    ('Batch Pulses', 'batchPulses'), ('Batch Interval', 'batchInterval'), ('Batch Temps', 'batchTemps'),
    ('SS Grind (Ode)', 'ssGrind'), ('Batch Grind (Ode)', 'batchGrind'),
    ('Validation', 'validation'), ('Latency', 'latency'),
]

ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 32
ws.column_dimensions['C'].width = 32

row = 1
# Title row
ws.merge_cells('A1:C1')
ws['A1'] = 'Aiden Recipe A/B Test: Claude Sonnet 4.6 vs GPT-5.4'
ws['A1'].font = Font(name='Arial', bold=True, size=14)
ws['A1'].alignment = Alignment(horizontal='center')
row = 3

for bean in beans:
    # Bean header
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws.cell(row=row, column=1, value=f"{bean['name']} — {bean['desc']}")
    cell.font = bean_font
    cell.fill = bean_fill
    cell.alignment = Alignment(horizontal='center')
    for c in range(1, 4):
        ws.cell(row=row, column=c).fill = bean_fill
        ws.cell(row=row, column=c).border = thin_border
    row += 1

    # Model headers
    ws.cell(row=row, column=1, value='Parameter').font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')

    ws.cell(row=row, column=2, value='Claude Sonnet 4.6').font = header_font
    ws.cell(row=row, column=2).fill = header_fill
    ws.cell(row=row, column=2).border = thin_border
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')

    ws.cell(row=row, column=3, value='GPT-5.4').font = header_font
    ws.cell(row=row, column=3).fill = header_fill
    ws.cell(row=row, column=3).border = thin_border
    ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
    row += 1

    # Data rows
    for i, (label, key) in enumerate(fields):
        bg = alt_row if i % 2 == 0 else white
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=1).fill = bg
        ws.cell(row=row, column=1).border = thin_border

        for col, model_key, model_fill_color in [(2, 'sonnet', sonnet_fill), (3, 'gpt', gpt_fill)]:
            val = bean[model_key][key]
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if key == 'validation':
                cell.fill = pass_fill if val == 'PASS' else warn_fill
                cell.font = Font(name='Arial', bold=True, size=10)
            elif key == 'bloomRatio' and isinstance(val, (int, float)) and val > 3:
                cell.fill = warn_fill
                cell.font = Font(name='Arial', bold=True, size=10, color='C0392B')
            else:
                cell.fill = bg
        row += 1

    row += 1  # Gap between beans

wb.save('/Users/talmeltzer/Documents/VIBE CODING/Coffee App Build /docs/data/aiden-ab-test-results.xlsx')
print('Saved docs/data/aiden-ab-test-results.xlsx')
